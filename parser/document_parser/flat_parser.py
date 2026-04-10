"""flat 포맷 파싱 - .md / .csv / .xlsx"""
import csv
from pathlib import Path

from .models import Block
from .llm import describe_table


def parse(file_path: str) -> list[Block]:
    ext = Path(file_path).suffix.lower()

    if ext == ".md":
        return _parse_md(file_path)
    elif ext == ".csv":
        return _parse_csv(file_path)
    elif ext == ".xlsx":
        return _parse_xlsx(file_path)
    else:
        raise ValueError(f"flat_parser 지원 포맷이 아닙니다: {ext}")


def _parse_md(path: str) -> list[Block]:
    text = Path(path).read_text(encoding="utf-8")
    return [Block(block_type="text", content=text, page=0, bbox=None)]


def _parse_csv(path: str) -> list[Block]:
    import chardet
    raw = Path(path).read_bytes()
    detected = chardet.detect(raw)
    encoding = detected.get("encoding") or "utf-8-sig"

    rows = []
    with open(path, encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for row in reader:
            rows.append(dict(row))

    if not rows:
        return []

    md_lines = ["| " + " | ".join(str(h) for h in headers) + " |"]
    md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        md_lines.append("| " + " | ".join(str(row.get(h, "")) for h in headers) + " |")
    table_md = "\n".join(md_lines)

    description = describe_table(table_md)
    file_name = Path(path).stem
    return [Block(
        block_type="table",
        content=f"[표 설명]\n{description}\n\n[표 원문]\n{table_md}",
        page=0,
        bbox=None,
        table_json=rows,
        sheet_name=file_name,
        header_depth=1,
        description=description,
    )]


def _expand_merged_cells(ws) -> None:
    """병합 셀 범위를 상단 좌측 값으로 채운 뒤 병합 해제."""
    for merge_range in list(ws.merged_cells.ranges):
        min_row = merge_range.min_row
        min_col = merge_range.min_col
        max_row = merge_range.max_row
        max_col = merge_range.max_col
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merge_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).value = top_left_value


def _detect_tables(ws) -> list[dict]:
    """완전히 빈 행을 경계로 독립 테이블 영역(min_row, max_row, min_col, max_col)을 분리."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    if max_row == 0 or max_col == 0:
        return []

    # 행별 비어있음 여부
    row_empty = []
    for r in range(1, max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        row_empty.append(all(v is None or str(v).strip() == "" for v in values))

    # 연속된 비어있지 않은 행 → 테이블 영역
    tables = []
    in_block = False
    block_start = None
    for r in range(1, max_row + 1):
        if not row_empty[r - 1]:
            if not in_block:
                in_block = True
                block_start = r
        else:
            if in_block:
                tables.append({
                    "min_row": block_start, "max_row": r - 1,
                    "min_col": 1, "max_col": max_col,
                })
                in_block = False
    if in_block:
        tables.append({
            "min_row": block_start, "max_row": max_row,
            "min_col": 1, "max_col": max_col,
        })
    return tables


def _detect_header_depth(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> int:
    """상단 N행이 전부 문자열(또는 None)이고 수직 병합 흔적이 있으면 다중 헤더로 판정."""
    header_depth = 0
    for r in range(min_row, max_row + 1):
        values = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
        # 한 셀이라도 수치·날짜 등 비문자 값이 있으면 헤더 행 종료
        if any(v is not None and not isinstance(v, str) for v in values):
            break
        if any(isinstance(v, str) and v.strip() for v in values):
            header_depth += 1
        else:
            break
    return max(header_depth, 1)


def _build_flat_headers(ws, min_row: int, header_depth: int, min_col: int, max_col: int) -> list[str]:
    """다중 헤더를 '부모.자식' 형태로 flat화."""
    if header_depth == 1:
        return [
            str(ws.cell(min_row, c).value).strip()
            if ws.cell(min_row, c).value is not None
            else f"col{c - min_col}"
            for c in range(min_col, max_col + 1)
        ]

    headers = []
    for c in range(min_col, max_col + 1):
        parts = []
        for r in range(min_row, min_row + header_depth):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        headers.append(".".join(parts) if parts else f"col{c - min_col}")

    # 중복 컬럼명 처리
    seen: dict[str, int] = {}
    deduped = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            deduped.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            deduped.append(h)
    return deduped


def _parse_xlsx(path: str) -> list[Block]:
    import openpyxl
    blocks: list[Block] = []
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)

    for sheet_idx, ws in enumerate(wb.worksheets):
        _expand_merged_cells(ws)
        tables = _detect_tables(ws)

        if not tables:
            continue

        for table_region in tables:
            min_row = table_region["min_row"]
            max_row = table_region["max_row"]
            min_col = table_region["min_col"]
            max_col = table_region["max_col"]

            header_depth = _detect_header_depth(ws, min_row, max_row, min_col, max_col)
            headers = _build_flat_headers(ws, min_row, header_depth, min_col, max_col)

            rows = []
            for r in range(min_row + header_depth, max_row + 1):
                row_data = {}
                for c_idx, c in enumerate(range(min_col, max_col + 1)):
                    v = ws.cell(r, c).value
                    row_data[headers[c_idx]] = str(v) if v is not None else ""
                rows.append(row_data)

            if not rows:
                continue

            md_lines = ["| " + " | ".join(headers) + " |"]
            md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            for row in rows:
                md_lines.append("| " + " | ".join(row.get(h, "") for h in headers) + " |")
            table_md = "\n".join(md_lines)

            description = describe_table(table_md)
            blocks.append(Block(
                block_type="table",
                content=f"[시트: {ws.title}]\n[표 설명]\n{description}\n\n[표 원문]\n{table_md}",
                page=sheet_idx,
                bbox=None,
                table_json=rows,
                sheet_name=ws.title,
                header_depth=header_depth,
                description=description,
            ))

    wb.close()
    return blocks
